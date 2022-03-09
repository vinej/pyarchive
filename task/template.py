from task.memory import Memory
from task.util import get_dict_value
import logging
from message.message import gmsg
import sys
import openpyxl
from copy import copy
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter

'''
The Template class is used to output memory info into an Excel template

The json object properties

Name        :   the name of the task
Kind        :   template
Description :   the description of the task
File        :   the excel output file
Template    :   the excel template file to use
'''
class Template:
    def __init__(self, jsondata):
        self.name = get_dict_value(jsondata,'Name')
        self.kind = get_dict_value(jsondata,'Kind')
        self.description = get_dict_value(jsondata,'Description')
        self.file = get_dict_value(jsondata,'File')
        self.template = get_dict_value(jsondata,'Template')
    #def


    def validate(self, mapcon, position):  
        _ = mapcon # not use here
        if self.name == None:
            logging.fatal(gmsg.get(26), position, 'Name')
            sys.exit(26)
        #if
        self.name = self.name.lower()

        if self.kind == None:
            logging.fatal(gmsg.get(27), position, self.name, 'Kind')
            sys.exit(27)
        #
        self.kind = self.kind.lower()
	#def

    def copyRange(self, startCol, startRow, endCol, endRow, sheet):
        rangeSelected = []
        rangeSelectedStyle = []
        #Loops through selected Rows
        for irow in range(startRow,endRow + 1,1):
            #Appends the row to a RowSelected list
            rowSelected = []
            rowSelectedStyle = []
            for jcol in range(startCol,endCol+1,1):
                rowSelected.append(sheet.cell(row = irow, column = jcol).value)
                rowSelectedStyle.append(copy(sheet.cell(row = irow, column = jcol)._style))
            #Adds the RowSelected List and nests inside the rangeSelected
            rangeSelected.append(rowSelected)
            rangeSelectedStyle.append(rowSelectedStyle)
        return rangeSelected, rangeSelectedStyle
    #def



    #Paste data from copyRange into template sheet
    def pasteRange(self, startCol, startRow, endCol, endRow, sheetReceiving, copiedData, row_ori):
        countRow = 0
        for irow in range(startRow,endRow+1,1):
            countCol = 0
            for jcol in range(startCol,endCol+1,1):
                value = copiedData[countRow][countCol]
                if str(value)[0] == '=':
                    ori = get_column_letter(jcol) + str( (row_ori+1) )
                    dst = get_column_letter(jcol) + str(irow)
                    sheetReceiving.cell(row = irow, column = jcol).value = Translator(value, origin=ori).translate_formula(dst)
                else:
                    sheetReceiving.cell(row = irow, column = jcol).value = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1
    #def

    #Paste data from copyRange into template sheet
    def pasteRangeStyle(self, startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
        countRow = 0
        for irow in range(startRow,endRow+1,1):
            countCol = 0
            for jcol in range(startCol,endCol+1,1):
                sheetReceiving.cell(row = irow, column = jcol)._style = copiedData[countRow][countCol]
                countCol += 1
            countRow += 1
    #def

    def find_cr(self, sheet, value):
        for row in range(sheet.max_row):
            for col in range(sheet.max_column):
                if sheet.cell(row+1,col+1).value == value:
                    print(col+1, row+1)
                    return (col+1,row+1)
        raise Exception('not found')
    #def

    '''
    Get a list of fields name from the cell with pattern {{xxx}}
    '''
    def get_field_name(self, value):
        outnames = []
        # return a list of name {{xxx}}
        list = str(value).split('{{')
        for part in list:
            idx = part.find('}}')
            if idx != -1:
                outnames.append('{{'+part[:idx+2])
        return outnames
    #def

    def replace_field_name(self, row, name, value):
        field = name.replace('{{','')
        field = field.replace('}}','')
        return value.replace(name, row[field])
    #def

    def replace_fields(self, row, startCol, startRow, endCol, endRow, sheet):
        #Loops through selected Rows
        for irow in range(startRow,endRow + 1,1):
            for jcol in range(startCol,endCol+1,1):
                if str(sheet.cell(irow,jcol).value).startswith('=') == False:
                    listnames = self.get_field_name(sheet.cell(irow,jcol).value)
                    for name in listnames:
                        sheet.cell(irow,jcol).value = self.replace_field_name(row, name, sheet.cell(irow,jcol).value)
    #def

    def run(self, mapmem, mapref, mapcon, position):
        logging.info(gmsg.get(4), self.kind, self.name)

        book = openpyxl.load_workbook(self.template)

        for sheet in book.worksheets:
            source = sheet.title
            cstart,rstart = self.find_cr(sheet, '{{begin}}')
            cend,rend = self.find_cr(sheet, '{{end}}')
            # the dynamic sedction could be on many rows. We need the height of the section
            height = rend - rstart - 1
            # 1,4
            # insert rows before the end. number of row into the data * the heiht of the section
            sheet.insert_rows(rend,(len(mapmem[source].rows) * height - 1))

            # copy the section into memory                                                 
            rng,sty = self.copyRange(cstart,rstart+1,sheet.max_column,rend-1+height-1,sheet)

            # paste the section into the added rows
            for i in range(len(mapmem[source].rows)):
                hstart = rstart + (i*height) + 1
                self.pasteRange(cstart,hstart,sheet.max_column,hstart + height - 1,sheet,rng,rstart+1)
                self.pasteRangeStyle(cstart,hstart,sheet.max_column,hstart + height - 1,sheet,sty)

            # update the data of each section
            for i in range(len(mapmem[source].rows)):
                hstart = rstart + (i*height) + 1
                self.replace_fields(mapmem[source].rows[i], cstart, hstart, sheet.max_column, hstart + height - 1, sheet)

            # remove the {{begin}} marker
            sheet.delete_rows(rstart,1)

            # remove the {{end}} marker
            sheet.delete_rows(rend+((len(mapmem[source].rows))*height)-2)
                
        book.save(self.file)

        logging.info(gmsg.get(3), self.kind, self.name)
    #def



#class

