import csv
import sys
from openpyxl import load_workbook

from output.memory import Memory
from output.util import get_dict_value, replace_global_parameter
import logging
from message.message import gmsg
from task.base import BaseTask
import sys

'''
The Excel class is used to create csv file from a complex excel file

The json object properties

Name        :   the name of the task
Kind        :   excel
Description :   the description of the task
Source      :   the excel source file
Output      :   the csv output file
Header      :   the csv header, comma separated
Actions     :   the file containing the list of actions to perform  
'''
class Excel(BaseTask):
    def __init__(self, jsondata):
        self.name = get_dict_value(jsondata,'Name')
        self.kind = get_dict_value(jsondata,'Kind')
        self.description = get_dict_value(jsondata,'Description')
        self.source = get_dict_value(jsondata,'Source')
        self.output = get_dict_value(jsondata,'Output')
        self.header =  get_dict_value(jsondata,'Header')
        self.actions = get_dict_value(jsondata,'Actions')
    #def

    def validate(self, mapcon, position):  
        _ = mapcon # not use here
        if self.name == None:
            logging.fatal(gmsg.get(26), position, 'Name')
            sys.exit(26)
        #if
        self.name = self.name.lower()

        if self.kind == None:
            logging.fatal(gmsg.get(27), position, self.name, 'Kind')
            sys.exit(27)
        #
        self.kind = self.kind.lower()
	#def

    def run(self, mapmem, mapref, mapcon, position, g_rows):
        
        self.description = replace_global_parameter(self.description, g_rows)
        self.source = replace_global_parameter(self.source, g_rows)
        self.output = replace_global_parameter(self.output, g_rows)
        self.header = replace_global_parameter(self.header, g_rows)
        self.actions = replace_global_parameter(self.actions, g_rows)

        logging.info(gmsg.get(4), self.kind, self.name)
        _ = mapcon    # not used for now
        _ = position  # not used for now
        _ = mapref # not used for now
        _ = g_rows

        # Load workbook and select the active sheet
        print('loading excel')
        wb = load_workbook(self.source, data_only=True)
        sheet = wb.active
        print('loaded')

        # Parse header
        headers = [h.strip() for h in self.header.split(',')]

        # Initialize position
        row_idx, col_idx = 1, 1  # Excel is 1-indexed

        # Initialize current row data
        current_row = {}
        output_rows = []

        # Read actions
        with open(self.actions, 'r') as f:
            actions = [line.strip() for line in f if line.strip()]
        #with

        action_idx = -1
        max_action = len(actions)
        labels = {}

        while action_idx < max_action-1 :
            action_idx += 1
            action = actions[action_idx]
            if ',' in action:
                cmd, arg = action.split(',', 1)
                cmd = cmd.strip().lower()
                arg = arg.strip()
            else:
                cmd = action.strip().lower()
                arg = None

            if cmd == 'label':
                labels[arg] = action_idx
        #while
        cnt = 0
        action_idx = -1
        while action_idx < max_action-1 :
            action_idx += 1
            cnt = cnt + 1
            if cnt == 10000:
                print(f'idx = {action_idx}, row = {row_idx}')
                cnt = 0
            #if
            action = actions[action_idx]
            if ',' in action:
                cmd, arg = action.split(',', 1)
                cmd = cmd.strip().lower()
                arg = arg.strip()
            else:
                cmd = action.strip().lower()
                arg = None
            #if

            if cmd == 'label':
                # do nothing
                continue
            elif cmd == 'sheet':
                sheet = wb[arg]
            elif cmd == 'cmp':
                # do nothing
                if "|" not in arg :
                    raise ValueError(f"if command requires '|' in argument")
                #if
                arg, sgoto = arg.split('|',1)
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    cell_value = str(cell_value).strip()  
                else:
                    cell_value = ''  
                #if
                if str(current_row[arg]) == cell_value :
                    action_idx = int(labels[sgoto])
                #if
            elif cmd == 'goto':
                action_idx = int(labels[arg])
            elif cmd == 'down':
                row_idx += int(arg)
            elif cmd == 'up':
                row_idx -= int(arg)
            elif cmd == 'right':
                col_idx += int(arg)                
            elif cmd == 'left':
                col_idx -= int(arg)
            elif cmd == 'create':
                dictcol = {}
                columns = []
                if arg is not None:
                    columns = arg.split('|')
                    for c in columns:
                        dictcol[c] = current_row.get(c,'') 
                    #for
                #if
                current_row = {col: '' for col in headers}
                if arg is not None:
                    for c in columns:
                        current_row[c] = dictcol[c] 
                    #for
                #if
            elif cmd == 'put':
                idx = -1
                sep = ''
                if '|' in arg :
                    arg,sep,idx = arg.split('|',2)  

                if arg not in headers:
                    raise ValueError(f"Column '{arg}' not in header")
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value is None or cell_value == '':
                    #do nothing for a None put
                    continue

                # arg could be only the name of the column or a split of the value
                if idx != -1:
                    try:
                        real_value = cell_value.split(sep)[int(idx)]
                    except:
                        real_value = cell_value

                    current_row[arg] = real_value
                else:
                    current_row[arg] = cell_value
                #if
            elif cmd == 'add':
                if arg not in headers:
                    raise ValueError(f"Column '{arg}' not in header")
                
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if current_row[arg] is None or current_row[arg] == '':
                    current_row[arg] = cell_value
                else:
                    try :
                        current_row[arg] =  float(current_row[arg]) + float(cell_value)
                    except:
                        current_row[arg] = cell_value
                    #try
                #if
            elif cmd == 'set':
                if '|' in arg :
                    arg,val = arg.split('|',1)  
                else:
                    raise ValueError(f"set command requires '|' in argument")
                #if
                if arg not in headers:
                    raise ValueError(f"Column '{arg}' not in header")
                
                current_row[arg] = val
            elif cmd == 'save':
                output_rows.append(current_row.copy())
            elif cmd == 'end':
                #output_rows.append(current_row.copy())
                break
            elif cmd == 'if':
                if "|" not in arg :
                    raise ValueError(f"if command requires '|' in argument")
                #if
                cmp, sgoto = arg.split('|',1)
                cell_value = sheet.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    cell_value = str(cell_value).strip()
                else:
                    cell_value = ''
                #if

                if cmp == 'empty' :
                    cmp = ''
                #if
                if cell_value == cmp :
                    action_idx = int(labels[sgoto])
                #if
            else:
                raise ValueError(f"Unknown command: {cmd}")
        #while

        # Write to CSV
        with open(self.output, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(output_rows)
    #def
#class
