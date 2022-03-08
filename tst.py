from copy import copy

def copyRange(startCol, startRow, endCol, endRow, sheet):
    rangeSelected = []
    rangeSelectedStyle = []
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        #Appends the row to a RowSelected list
        rowSelected = []
        rowSelectedStyle = []
        for j in range(startCol,endCol+1,1):
            rowSelected.append(sheet.cell(row = i, column = j).value)
            rowSelectedStyle.append(copy(sheet.cell(row = i, column = j)._style))
        #Adds the RowSelected List and nests inside the rangeSelected
        rangeSelected.append(rowSelected)
        rangeSelectedStyle.append(rowSelectedStyle)
    return rangeSelected, rangeSelectedStyle
#def

#Paste data from copyRange into template sheet
def pasteRange(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            sheetReceiving.cell(row = i, column = j).value = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1
#def

#Paste data from copyRange into template sheet
def pasteRangeStyle(startCol, startRow, endCol, endRow, sheetReceiving, copiedData):
    countRow = 0
    for i in range(startRow,endRow+1,1):
        countCol = 0
        for j in range(startCol,endCol+1,1):
            sheetReceiving.cell(row = i, column = j)._style = copiedData[countRow][countCol]
            countCol += 1
        countRow += 1
#def

def find_cr(sheet, value):
    for r in range(sheet.max_row):
        for c in range(sheet.max_column):
            if sheet.cell(r+1,c+1).value == value:
                print(c+1, r+1)
                return (c+1,r+1)
    raise Exception('not found')
#def

def replace_fields(startCol, startRow, endCol, endRow, sheet):
    #Loops through selected Rows
    for i in range(startRow,endRow + 1,1):
        for j in range(startCol,endCol+1,1):
            pass
        #Adds the RowSelected List and nests inside the rangeSelected
#def

import openpyxl
book = openpyxl.load_workbook('C:/Users/jyvin/OneDrive/Documents/GitHub/pyarchive/template.xlsx')

for sheet in book.worksheets:
    c1,r1 = find_cr(sheet, '{{begin}}')
    # 1,2
    c2,r2 = find_cr(sheet, '{{end}}')
    # 1,4
    # insert rows before the end
    sheet.insert_rows(r2,10-1)
    rng,sty = copyRange(c1,r1+1,sheet.max_column,r2-1,sheet)
    for i in range(10):
        pasteRange(c1,r1+i+1,sheet.max_column,r1+i+1,sheet,rng)
        pasteRangeStyle(c1,r1+i+1,sheet.max_column,r1+i+1,sheet,sty)
        
book.save('C:/Users/jyvin/OneDrive/Documents/GitHub/pyarchive/out_template.xlsx')
